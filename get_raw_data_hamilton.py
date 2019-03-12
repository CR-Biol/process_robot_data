
import os
import logging

import xlrd
from openpyxl import Workbook as openpyxlWorkbook

import constants


# Initialize logger.
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG) # DEBUG, INFO, WARNING, ERROR, CRITICAL
formatter = constants.LOG_FORMATTER
if not os.path.exists("logs"):
    os.makedirs("logs")
log_file_handler = logging.FileHandler(os.path.join("logs", __name__ + ".log"))
log_file_handler.setFormatter(formatter)
logger.addHandler(log_file_handler)


#==============================================================================
# READ DATA FROM HAMILTON ROBOT AS EXCEL FILE
#==============================================================================

def read_raw_data(reporter_name, appendix="_results.txt"):
    BARCODE = "THIS-IS-A-GENERIC-PLACEHOLDER-FOR-ANY-BARCODE-IN-DATA-FILES" 
    written_barcodes = [] # User feedback on GUI; function return value.
    data_files = [file for file in os.listdir() if (".xls" in file) and
                                                   ("P4" in file)]
    cyc_min, cyc_max = float("inf"), -float("inf") # Initializing variables for measurement cycles
    first_iteration = True

    od = constants.header
    od_list = []
    fu = constants.header
    fu_list = []
    for data_file in data_files:
        # Iterating through all data files to gain info about barcode, basic
        #  file name (which can be different between runs), and maximal
        #  measurement cycle.
        barcode = data_file.split(sep="_")[0].replace("BC", "")
        if not barcode in written_barcodes:
            written_barcodes.append(barcode)

        if first_iteration:
            basic_file_name = "_".join(data_file.replace(barcode, BARCODE).split(sep="_")[:-1])
            first_iteration = False

        cyc_current = int(data_file.split(sep="_")[-1].split(sep=".")[0])
        if cyc_current < cyc_min:
            cyc_min = cyc_current
        elif cyc_current > cyc_max:
            cyc_max = cyc_current

    for barcode in written_barcodes:
        od = "OD600\n" + constants.header
        fu = reporter_name + "\n" + constants.header
        for cyc_num in range(cyc_min, cyc_max+1):
            current_file = basic_file_name.replace(BARCODE, barcode) + \
                                                    "_{}.xls".format(cyc_num)
            wb = xls_to_xlsx_conversion(current_file)
            ws = wb.active
            curr_time = ws["H2"].value # Time info is stored as a float representing days since 1900.

            try:
                td = curr_time - first_time
            except NameError:
                first_time = curr_time
                td = 0
            time_diff_in_min = round(24 * 60 * td)

            temperature = "0" # Since temperature is later converted to a
                              #  numeric value, NA or ND would cause errors
            od += "\t".join([str(cyc_num), str(time_diff_in_min), str(temperature)]) + "\t"
            fu += "\t".join([str(cyc_num), str(time_diff_in_min), str(temperature)]) + "\t"

            for i in range(2, 97 + 1): # Cells C2 to C97 and D2 to D97 contain data
                od += str(ws["C{}".format(i)].value).replace(",", ".") + "\t"
                fu += str(ws["D{}".format(i)].value).replace(",", ".") + "\t"

            od += "\n"
            fu += "\n"

        od += "\n" # Seperates OD and reporter

        # Writing output/results TSV
        output_file = barcode + appendix
        with open(output_file, "w") as f:
            f.write(od)
            f.write(fu)

    return written_barcodes


def xls_to_xlsx_conversion(file):
    """Converts .xls to .xlsx files.
    Function taken with minor adjustments from GitHub user malexandre,
    https://gist.github.com/malexandre/730223fc089f70c65a7d"""
    xlsBook = xlrd.open_workbook(filename=file)
    workbook = openpyxlWorkbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)
    return workbook
