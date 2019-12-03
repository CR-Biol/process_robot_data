
import os
import logging
import datetime

from openpyxl import load_workbook

import constants


# Initialize logger.
logger = constants.setup_logger(
    log_level = logging.DEBUG,
    logger_name = __name__
    )



#==============================================================================
# READ DATA FROM PLATE READER AS EXCEL FILE
#==============================================================================

def time_from_excel(cell_value):
    """Takes the value from a Excel sheet cell containing information about time.
    Returns a datetime-object with:
    (year, month, day, hour, min, sec)

    Requires datetime module.
    """
    day_info, time_info = cell_value.split()
    day, month, year = day_info.split(sep=".")
    day, month, year = int(day), int(month), int(year)
    hour, min, second = time_info.split(sep=":")
    hour, min, second = int(hour), int(min), int(second)
    return datetime.datetime(year, month, day, hour, min, second)


def read_raw_data(appendix="_results.txt", reporter_name="this-is-a-dummy-variable"):
    """Reads raw data from excel file.
    Assumes only ONE raw data file containing several data sheets. Those sheets
    should be named "SheetX" where X is the measurement cycle.

    State: 28/07/2018"""

    all_files = [file for file in os.listdir() if file.endswith("xlsx")]
    written_barcodes = [] # List for user feedback

    is_first_file = False

    for file in all_files:
        to_write = ""
        barcode = file.split(sep=".")[0]
        data_wb = load_workbook(file)
        first_timepoint = "not_defined"
        valid_sheets = [sheet for sheet in data_wb.sheetnames if "Sheet" in sheet]
        valid_sheets.sort()
        number_of_measurements = len(valid_sheets)

        ws = data_wb[valid_sheets[0]] # Assign first worksheet.

        length_of_sheet = 0

        for _ in ws.rows: # Getting length of data entries in the Excel sheet.
            length_of_sheet += 1

        for row_num in range(1, length_of_sheet+1):
            # Iterating through every row of the first column
            curr_cell = "A" + str(row_num)

            if ws[curr_cell].value in ("Well", "<>"): # Detected entry with data.
                # Label Handling. Finding names for reporters.
                label_found = False
                curr_row = int(row_num)
                while not label_found:
                    potential_label_cell = "A" + str(curr_row - 1)
                    if ws[potential_label_cell].value is not None:
                        if "Label" in ws[potential_label_cell].value:
                            label_found = True
                            label = ws[potential_label_cell].value.split(": ")[1]
                        else:
                            curr_row -= 1
                    else:
                        curr_row -= 1

                # Get first time point and locate cell with time information.
                #  Read only ONE time point per sheet!
                curr_row = int(row_num)
                while first_timepoint == "not_defined":
                    potential_time_cell = "A" + str(curr_row - 1)
                    if ws[potential_time_cell].value is not None:
                        if "Start Time" in ws[potential_time_cell].value:
                            first_timepoint = time_from_excel(ws["B"+str(curr_row-1)].value)
                            time_cell = "B" + potential_time_cell[1:]
                        else:
                            curr_row -= 1
                    else:
                        curr_row -= 1

                # Temperature is not recorded in the Excel files.
                temperature = "-1" # Since temperature is later converted to a
                                  #  numeric value, NA or ND would cause errors

                # Print headline
                to_write += label + "\n"
                to_write += constants.header

                # Data management.
                for sheet in valid_sheets:
                    ws = data_wb[sheet]
                    data_names = []
                    time_of_measurement = time_from_excel(ws[time_cell].value)
                    time_in_seconds = (time_of_measurement - first_timepoint).total_seconds()
                    time_in_min = time_in_seconds / 60
                    data_vals = [str(round(time_in_min, 2)), temperature]
                    cell_with_data = "B" + str(row_num+1) # First cell containing data.
                    while not ws[cell_with_data].value is None: # Iterating until end of data entry.
                        curr_row = cell_with_data[1:]
                        data_names.append(ws["A"+curr_row].value)
                        data_vals.append(str(round(ws["B"+curr_row].value, 3)))
                        cell_with_data = "B" + str(int(cell_with_data[1:]) + 1)
                    data_vals.insert(0, sheet[-1])
                    to_write += "\t".join(data_vals) + "\n"
                to_write += "\n"
                outfile = barcode + appendix
                with open(outfile, "w") as out:
                    out.write(to_write)

        written_barcodes.append(barcode)
        ws = data_wb[valid_sheets[0]]
        is_first_file = False

    return written_barcodes
